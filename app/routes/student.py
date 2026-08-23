"""
Student routes blueprint.

This module provides web routes for managing students through the Flask interface.
"""

import csv
import io
import logging
import os
from pathlib import Path
from typing import Any, cast
from uuid import uuid4

from flask import (
    Blueprint,
    Response,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import login_required
from sqlalchemy.exc import SQLAlchemyError
from werkzeug.datastructures import FileStorage

from app import db
from app.forms.student import StudentForm, StudentImportForm
from app.models.course import Course
from app.models.enrollment import Enrollment
from app.models.grade import Grade
from app.models.student import Student, validate_email
from app.models.submission import Submission
from app.services.student_service import StudentService
from app.utils.auth import admin_required
from app.utils.pagination import paginate_query

# Configure logging
logger = logging.getLogger(__name__)

# Create blueprint
bp = Blueprint("student", __name__, url_prefix="/students")

REQUIRED_IMPORT_HEADERS = [
    "first_name",
    "last_name",
    "student_id",
    "email",
    "program",
]


def _normalize_header(header: str | None) -> str:
    if not header:
        return ""
    return header.strip().lower()


def _load_csv_rows(file: FileStorage) -> tuple[list[str], list[dict[str, str]]]:
    file.stream.seek(0)
    text_stream = io.TextIOWrapper(file.stream, encoding="utf-8-sig")
    reader = csv.DictReader(text_stream)
    raw_headers = reader.fieldnames or []
    normalized_headers = [_normalize_header(h) for h in raw_headers]
    header_map = dict(zip(raw_headers, normalized_headers, strict=False))

    rows: list[dict[str, str]] = []
    for row in reader:
        normalized_row: dict[str, str] = {}
        for raw_header, value in row.items():
            normalized_key = header_map.get(raw_header, "")
            if not normalized_key:
                continue
            normalized_row[normalized_key] = str(value).strip() if value else ""
        rows.append(normalized_row)
    return normalized_headers, rows


def _load_xlsx_rows(file: FileStorage) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "openpyxl is required for XLSX import. Install it and try again."
        ) from exc

    file.stream.seek(0)
    workbook = load_workbook(
        filename=io.BytesIO(file.stream.read()),
        read_only=True,
        data_only=True,
    )
    sheet = workbook.active
    if sheet is None:
        raise ValueError("XLSX file has no active sheet.")

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise ValueError("XLSX file has no rows.") from None

    raw_headers = [_normalize_header(str(h)) for h in headers]

    rows: list[dict[str, str]] = []
    for row in rows_iter:
        normalized_row: dict[str, str] = {}
        for idx, value in enumerate(row):
            header = raw_headers[idx] if idx < len(raw_headers) else ""
            if not header:
                continue
            normalized_row[header] = str(value).strip() if value is not None else ""
        rows.append(normalized_row)
    return raw_headers, rows


def _load_xls_rows(file: FileStorage) -> tuple[list[str], list[dict[str, str]]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError(
            "xlrd is required for XLS import. Install it and try again."
        ) from exc

    file.stream.seek(0)
    workbook = xlrd.open_workbook(file_contents=file.stream.read())
    if workbook.nsheets == 0:
        raise ValueError("XLS file has no sheets.")

    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("XLS file has no rows.")

    raw_headers = [_normalize_header(str(h)) for h in sheet.row_values(0)]

    rows: list[dict[str, str]] = []
    for row_idx in range(1, sheet.nrows):
        row = sheet.row_values(row_idx)
        normalized_row: dict[str, str] = {}
        for idx, value in enumerate(row):
            header = raw_headers[idx] if idx < len(raw_headers) else ""
            if not header:
                continue
            normalized_row[header] = str(value).strip() if value else ""
        rows.append(normalized_row)
    return raw_headers, rows


def _load_import_headers(file: FileStorage, fmt: str | None) -> list[tuple[str, str]]:
    format_hint = fmt or file.filename.rsplit(".", 1)[-1].lower()
    if format_hint == "csv":
        file.stream.seek(0)
        content = file.stream.read()
        file.stream.seek(0)
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        try:
            raw_headers = next(reader)
        except StopIteration:
            raise ValueError("CSV file has no rows.") from None
        normalized_headers = [_normalize_header(h) for h in raw_headers]
        return [
            (normalized_headers[idx], str(raw_header).strip())
            for idx, raw_header in enumerate(raw_headers)
            if normalized_headers[idx]
        ]
    if format_hint == "xlsx":
        file.stream.seek(0)
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError(
                "openpyxl is required for XLSX import. Install it and try again."
            ) from exc
        workbook = load_workbook(
            filename=io.BytesIO(file.stream.read()),
            read_only=True,
            data_only=True,
        )
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            raise ValueError("XLSX file has no rows.") from None
        raw_headers = [str(h).strip() if h is not None else "" for h in headers]
        normalized_headers = [_normalize_header(h) for h in raw_headers]
        return [
            (normalized_headers[idx], raw_header)
            for idx, raw_header in enumerate(raw_headers)
            if normalized_headers[idx]
        ]
    if format_hint == "xls":
        file.stream.seek(0)
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError(
                "xlrd is required for XLS import. Install it and try again."
            ) from exc
        workbook = xlrd.open_workbook(file_contents=file.stream.read())
        if workbook.nsheets == 0:
            raise ValueError("XLS file has no sheets.")
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            raise ValueError("XLS file has no rows.")
        raw_headers = [
            str(h).strip() if h is not None else "" for h in sheet.row_values(0)
        ]
        normalized_headers = [_normalize_header(h) for h in raw_headers]
        return [
            (normalized_headers[idx], raw_header)
            for idx, raw_header in enumerate(raw_headers)
            if normalized_headers[idx]
        ]
    raise ValueError(f"Unsupported format '{format_hint}'. Use csv, xlsx, or xls.")


def _load_import_rows(
    file: FileStorage, fmt: str | None, mapping: dict[str, str] | None = None
) -> list[dict[str, str]]:
    format_hint = fmt or file.filename.rsplit(".", 1)[-1].lower()
    if format_hint == "csv":
        headers, rows = _load_csv_rows(file)
    elif format_hint == "xlsx":
        headers, rows = _load_xlsx_rows(file)
    elif format_hint == "xls":
        headers, rows = _load_xls_rows(file)
    else:
        raise ValueError(f"Unsupported format '{format_hint}'. Use csv, xlsx, or xls.")

    if mapping:
        missing_map = [h for h in REQUIRED_IMPORT_HEADERS if not mapping.get(h)]
        if missing_map:
            raise ValueError("Mapping missing for: " + ", ".join(missing_map))
        used_headers = list(mapping.values())
        if len(set(used_headers)) != len(used_headers):
            raise ValueError("Mapping contains duplicate column selections.")
        unknown_headers = [h for h in used_headers if h not in headers]
        if unknown_headers:
            raise ValueError(
                "Mapping references unknown headers: " + ", ".join(unknown_headers)
            )
        mapped_rows: list[dict[str, str]] = []
        for row in rows:
            mapped_rows.append(
                {
                    key: row.get(mapping[key], "").strip()
                    for key in REQUIRED_IMPORT_HEADERS
                }
            )
        return mapped_rows

    missing = [h for h in REQUIRED_IMPORT_HEADERS if h not in headers]
    if missing:
        raise ValueError("Missing required headers: " + ", ".join(missing))
    return rows


def _save_import_file(file: FileStorage) -> tuple[str, Path, str]:
    filename = file.filename or "import"
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else "data"
    token = uuid4().hex
    import_dir = Path(current_app.root_path) / "uploads" / "imports"
    import_dir.mkdir(parents=True, exist_ok=True)
    file_path = import_dir / f"students_{token}.{extension}"
    file.save(file_path)
    return token, file_path, extension


@bp.route("/")
@login_required
def index() -> str:
    """
    List all students with optional search, program filter, and pagination.

    Query parameters:
        search: Optional search term to filter by name, student_id, or email
        program: Optional program filter
        page: Page number (default: 1)

    Returns:
        Rendered template with paginated list of students
    """
    search_term = request.args.get("search", "").strip()
    program_filter = request.args.get("program", "").strip()
    service = StudentService()

    try:
        # Build query using service's query method
        query = service.query(Student).filter(Student.deleted_at.is_(None))

        if search_term:
            search_pattern = f"%{search_term}%"
            query = query.filter(
                (Student.first_name.ilike(search_pattern))
                | (Student.last_name.ilike(search_pattern))
                | (Student.student_id.ilike(search_pattern))
                | (Student.email.ilike(search_pattern))
            )

        if program_filter:
            program_pattern = f"%{program_filter}%"
            query = query.filter(Student.program.ilike(program_pattern))

        query = query.order_by(Student.last_name, Student.first_name)
        pagination = paginate_query(query, per_page=20)

        return render_template(
            "student/list.html",
            students=pagination.items,
            pagination=pagination,
            search_term=search_term,
            program_filter=program_filter,
        )

    except SQLAlchemyError as e:
        logger.error(f"Database error while listing students: {e}")
        flash("Error loading students. Please try again.", "error")
        return render_template(
            "student/list.html",
            students=[],
            pagination=None,
            search_term="",
            program_filter="",
        )


@bp.route("/<int:student_id>")
@login_required
def show(student_id: int) -> str | Any:
    """
    Show details for a specific student.

    Args:
        student_id: Student database ID

    Returns:
        Rendered template with student details or redirect
    """
    service = StudentService()

    try:
        student = service.get_student(student_id)

        if not student:
            flash(f"Student mit Datenbank-ID {student_id} nicht gefunden.", "error")
            return redirect(url_for("student.index"))

        enrollments = (
            db.session.query(Enrollment)
            .join(Course)
            .filter(Enrollment.student_id == student.id)
            .order_by(Course.semester.desc(), Course.name)
            .all()
        )
        grades_count = (
            db.session.query(Grade)
            .join(Enrollment)
            .filter(Enrollment.student_id == student.id)
            .count()
        )
        submissions = (
            db.session.query(Submission)
            .join(Enrollment)
            .filter(Enrollment.student_id == student.id)
            .order_by(Submission.submission_date.desc())
            .limit(5)
            .all()
        )

        return render_template(
            "student/detail.html",
            student=student,
            enrollments=enrollments,
            grades_count=grades_count,
            submissions=submissions,
        )

    except SQLAlchemyError as e:
        logger.error(f"Database error while fetching student: {e}")
        flash("Error loading student details. Please try again.", "error")
        return redirect(url_for("student.index"))


@bp.route("/new", methods=["GET", "POST"])
@login_required
def new() -> str | Any:
    """
    Create a new student.

    GET: Show form
    POST: Create student and redirect to detail page

    Form fields:
        first_name: First name (required)
        last_name: Last name (required)
        student_id: Student ID (required, 8 digits)
        email: Email address (required)
        program: Study program (required)

    Returns:
        Rendered form template (GET) or redirect to detail page (POST)
    """
    form = StudentForm()
    service = StudentService()

    if form.validate_on_submit():
        try:
            # Create new student using service
            student = service.add_student(
                first_name=cast(str, form.first_name.data),
                last_name=cast(str, form.last_name.data),
                student_id=cast(str, form.student_id.data),
                email=cast(str, form.email.data),
                program=cast(str, form.program.data),
            )

            logger.info(
                f"Created student: {student.first_name} {student.last_name} ({student.student_id})"
            )
            flash(
                f"Student '{student.first_name} {student.last_name}' created successfully.",
                "success",
            )
            return redirect(url_for("student.show", student_id=student.id))

        except ValueError as e:
            logger.error(f"Validation error while creating student: {e}")
            flash(str(e), "error")

        except SQLAlchemyError as e:
            logger.error(f"Database error while creating student: {e}")
            flash("Error creating student. Please try again.", "error")

    # Display form validation errors
    for _field, errors in form.errors.items():
        for error in errors:
            flash(error, "error")

    return render_template("student/form.html", student=None, form=form)


@bp.route("/export")
@login_required
def export_students() -> Response | Any:
    """Export students to CSV with optional filters."""
    search_term = request.args.get("search", "").strip()
    program_filter = request.args.get("program", "").strip()
    service = StudentService()

    try:
        query = service.query(Student).filter(Student.deleted_at.is_(None))

        if search_term:
            search_pattern = f"%{search_term}%"
            query = query.filter(
                (Student.first_name.ilike(search_pattern))
                | (Student.last_name.ilike(search_pattern))
                | (Student.student_id.ilike(search_pattern))
                | (Student.email.ilike(search_pattern))
            )

        if program_filter:
            program_pattern = f"%{program_filter}%"
            query = query.filter(Student.program.ilike(program_pattern))

        students = query.order_by(Student.last_name, Student.first_name).all()
    except SQLAlchemyError as e:
        logger.error(f"Database error while exporting students: {e}")
        flash("Fehler beim Export der Studierenden.", "error")
        return redirect(url_for("student.index"))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["first_name", "last_name", "student_id", "email", "program"])
    for student in students:
        writer.writerow(
            [
                student.first_name,
                student.last_name,
                student.student_id,
                student.email,
                student.program,
            ]
        )

    filename = "students_export.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@bp.route("/import", methods=["GET", "POST"])
@login_required
@admin_required
def import_students() -> str | Any:
    """Import students from CSV/XLSX/XLS."""
    form = StudentImportForm()
    service = StudentService()
    import_errors: list[str] = []
    import_summary: dict[str, int] | None = None
    mapping_headers: list[tuple[str, str]] = []
    mapping_token = ""
    mapping_format = ""
    mapping_defaults: dict[str, str] = {}
    mapping_on_duplicate = ""

    if request.method == "POST" and request.form.get("mapping_token"):
        mapping_token = request.form.get("mapping_token", "").strip()
        mapping_format = request.form.get("file_format", "").strip()
        mapping_on_duplicate = request.form.get("on_duplicate", "skip").strip()
        file_extension = request.form.get("file_extension", "data").strip()
        file_path = (
            Path(current_app.root_path)
            / "uploads"
            / "imports"
            / f"students_{mapping_token}.{file_extension}"
        )

        mapping = {
            key: request.form.get(f"map_{key}", "").strip()
            for key in REQUIRED_IMPORT_HEADERS
        }
        mapping_defaults = mapping.copy()

        if not file_path.exists():
            flash("Importdatei nicht gefunden. Bitte erneut hochladen.", "error")
            return redirect(url_for("student.import_students"))

        try:
            with file_path.open("rb") as handle:
                file = FileStorage(stream=handle, filename=file_path.name)
                mapping_headers = _load_import_headers(file, mapping_format)
                handle.seek(0)
                rows = _load_import_rows(file, mapping_format, mapping=mapping)
        except ValueError as exc:
            import_errors.append(str(exc))
            return render_template(
                "student/import.html",
                form=form,
                import_errors=import_errors,
                mapping_headers=mapping_headers,
                mapping_token=mapping_token,
                mapping_format=mapping_format,
                mapping_defaults=mapping_defaults,
                mapping_on_duplicate=mapping_on_duplicate,
                file_extension=file_extension,
            )
    elif form.validate_on_submit():
        file = form.file.data
        try:
            token, file_path, extension = _save_import_file(file)
            with file_path.open("rb") as handle:
                stored_file = FileStorage(stream=handle, filename=file_path.name)
                mapping_headers = _load_import_headers(
                    stored_file, form.file_format.data
                )
            mapping_token = token
            mapping_format = form.file_format.data
            mapping_on_duplicate = form.on_duplicate.data
            normalized_headers = [value for value, _ in mapping_headers]
            mapping_defaults = {
                key: key if key in normalized_headers else ""
                for key in REQUIRED_IMPORT_HEADERS
            }
            return render_template(
                "student/import.html",
                form=form,
                import_errors=import_errors,
                mapping_headers=mapping_headers,
                mapping_token=mapping_token,
                mapping_format=mapping_format,
                mapping_defaults=mapping_defaults,
                mapping_on_duplicate=mapping_on_duplicate,
                file_extension=extension,
            )
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template(
                "student/import.html",
                form=form,
                import_errors=import_errors,
            )
    else:
        rows = []
        if request.method == "POST":
            for field_errors in form.errors.values():
                import_errors.extend(field_errors)

    if "rows" in locals() and not rows and mapping_token:
        flash("Keine Datenzeilen gefunden.", "warning")
        file_extension = request.form.get("file_extension", "data").strip()
        file_path = (
            Path(current_app.root_path)
            / "uploads"
            / "imports"
            / f"students_{mapping_token}.{file_extension}"
        )
        if file_path.exists():
            try:
                os.remove(file_path)
            except OSError:
                logger.warning("Failed to remove import file %s", file_path)
        return render_template(
            "student/import.html",
            form=form,
            import_errors=import_errors,
        )

    if "rows" in locals() and rows:
        created = updated = skipped = errors = 0
        seen_student_ids: set[str] = set()
        seen_emails: set[str] = set()

        for idx, row in enumerate(rows, start=2):
            missing_values = [
                header for header in REQUIRED_IMPORT_HEADERS if not row.get(header)
            ]
            if missing_values:
                errors += 1
                import_errors.append(
                    f"Zeile {idx}: fehlende Werte: {', '.join(missing_values)}"
                )
                continue

            student_id = row["student_id"].strip()
            email = row["email"].strip().lower()

            if not student_id:
                errors += 1
                import_errors.append(f"Zeile {idx}: fehlende Matrikelnummer")
                continue
            if not validate_email(email):
                errors += 1
                import_errors.append(f"Zeile {idx}: ungültige E-Mail {email}")
                continue

            if student_id in seen_student_ids or email in seen_emails:
                errors += 1
                import_errors.append(f"Zeile {idx}: Duplikat innerhalb der Datei")
                continue
            seen_student_ids.add(student_id)
            seen_emails.add(email)

            existing_by_id = (
                db.session.query(Student).filter_by(student_id=student_id).first()
            )
            existing_by_email = db.session.query(Student).filter_by(email=email).first()

            existing: Student | None = None
            if existing_by_id and existing_by_email:
                if existing_by_id.id != existing_by_email.id:
                    errors += 1
                    import_errors.append(
                        f"Zeile {idx}: Matrikelnummer und E-Mail gehören zu unterschiedlichen Datensätzen"
                    )
                    continue
                existing = existing_by_id
            else:
                existing = existing_by_id or existing_by_email

            if existing:
                if form.on_duplicate.data == "skip":
                    skipped += 1
                    continue
                if form.on_duplicate.data == "error":
                    errors += 1
                    import_errors.append(f"Zeile {idx}: Duplikat gefunden")
                    continue
                if form.on_duplicate.data == "update":
                    if existing.deleted_at:
                        existing.deleted_at = None  # type: ignore
                    try:
                        existing_id = int(existing.id)
                        updated_student = service.update_student(
                            existing_id,
                            row["first_name"],
                            row["last_name"],
                            student_id,
                            email,
                            row["program"],
                            validate_id=False,
                        )
                        if updated_student:
                            updated += 1
                        else:
                            errors += 1
                            import_errors.append(
                                f"Zeile {idx}: Datensatz nicht gefunden"
                            )
                    except ValueError as exc:
                        errors += 1
                        import_errors.append(f"Zeile {idx}: {exc}")
                    except SQLAlchemyError:
                        errors += 1
                        import_errors.append(f"Zeile {idx}: Datenbankfehler")
                    continue

            try:
                service.add_student(
                    row["first_name"],
                    row["last_name"],
                    student_id,
                    email,
                    row["program"],
                    validate_id=False,
                )
                created += 1
            except ValueError as exc:
                errors += 1
                import_errors.append(f"Zeile {idx}: {exc}")
            except SQLAlchemyError:
                errors += 1
                import_errors.append(f"Zeile {idx}: Datenbankfehler")

        flash(
            f"Import abgeschlossen. Neu: {created}, "
            f"Aktualisiert: {updated}, Übersprungen: {skipped}, "
            f"Fehler: {errors}",
            "success" if errors == 0 else "warning",
        )
        import_summary = {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
        }
        if mapping_token and mapping_format is not None:
            file_extension = request.form.get("file_extension", "data").strip()
            file_path = (
                Path(current_app.root_path)
                / "uploads"
                / "imports"
                / f"students_{mapping_token}.{file_extension}"
            )
            if file_path.exists():
                try:
                    os.remove(file_path)
                except OSError:
                    logger.warning("Failed to remove import file %s", file_path)

    return render_template(
        "student/import.html",
        form=form,
        import_errors=import_errors,
        import_summary=import_summary,
        mapping_headers=mapping_headers,
        mapping_token=mapping_token,
        mapping_format=mapping_format,
        mapping_defaults=mapping_defaults,
        mapping_on_duplicate=mapping_on_duplicate,
        file_extension=request.form.get("file_extension", "").strip(),
    )


@bp.route("/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
def edit(student_id: int) -> str | Any:
    """
    Edit an existing student.

    GET: Show edit form
    POST: Update student and redirect to detail page

    Args:
        student_id: Student database ID

    Form fields:
        first_name: First name (required)
        last_name: Last name (required)
        student_id: Student ID (required)
        email: Email address (required)
        program: Study program (required)

    Returns:
        Rendered form template (GET) or redirect to detail page (POST)
    """
    service = StudentService()

    try:
        student = service.get_student(student_id)

        if not student:
            flash(f"Student mit Datenbank-ID {student_id} nicht gefunden.", "error")
            return redirect(url_for("student.index"))

        form = StudentForm(student=student, obj=student)

        if form.validate_on_submit():
            try:
                # Update using service
                student = service.update_student(
                    student_id=student_id,
                    first_name=form.first_name.data,
                    last_name=form.last_name.data,
                    student_number=form.student_id.data,
                    email=form.email.data,
                    program=form.program.data,
                )

                if student:
                    logger.info(
                        f"Updated student: {student.first_name} {student.last_name} ({student.student_id})"
                    )
                    flash(
                        f"Student '{student.first_name} {student.last_name}' updated successfully.",
                        "success",
                    )
                    return redirect(url_for("student.show", student_id=student.id))

            except ValueError as e:
                logger.error(f"Validation error while updating student: {e}")
                flash(str(e), "error")

            except SQLAlchemyError as e:
                logger.error(f"Database error while updating student: {e}")
                flash("Error updating student. Please try again.", "error")

        # Display form validation errors
        for _field, errors in form.errors.items():
            for error in errors:
                flash(error, "error")

        return render_template("student/form.html", student=student, form=form)

    except SQLAlchemyError as e:
        logger.error(f"Database error while loading student: {e}")
        flash("Error loading student. Please try again.", "error")
        return redirect(url_for("student.index"))


@bp.route("/<int:student_id>/delete", methods=["GET", "POST"])
@login_required
@admin_required
def delete(student_id: int) -> str | Any:
    """
    Delete a student.

    GET: Show confirmation page
    POST: Delete student and redirect to list

    Args:
        student_id: Student database ID

    Returns:
        Rendered confirmation template (GET) or redirect to list (POST)
    """
    service = StudentService()

    try:
        student = service.get_student(student_id)

        if not student:
            flash(f"Student mit Datenbank-ID {student_id} nicht gefunden.", "error")
            return redirect(url_for("student.index"))

        if request.method == "GET":
            return render_template("student/delete.html", student=student)

        # POST: Delete student using service
        student_name = f"{student.first_name} {student.last_name}"
        if service.delete_student(student_id):
            flash(f"Student '{student_name}' deleted successfully.", "success")
            return redirect(url_for("student.index"))

        flash(f"Error deleting student '{student_name}'.", "error")
        return redirect(url_for("student.show", student_id=student_id))

    except SQLAlchemyError as e:
        logger.error(f"Database error while deleting student: {e}")
        flash("Error deleting student. Please try again.", "error")
        return redirect(url_for("student.show", student_id=student_id))
