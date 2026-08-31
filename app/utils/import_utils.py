"""Shared utilities for CSV/XLSX/XLS import across all entities."""

import csv
import io
import logging
from pathlib import Path
from uuid import uuid4

from flask import current_app
from werkzeug.datastructures import FileStorage

logger = logging.getLogger(__name__)


def normalize_header(header: str | None) -> str:
    if not header:
        return ""
    return header.strip().lower()


def _load_csv_rows(file: FileStorage) -> tuple[list[str], list[dict[str, str]]]:
    file.stream.seek(0)
    content = file.stream.read()
    file.stream.seek(0)
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    raw_headers = list(reader.fieldnames or [])
    normalized_headers = [normalize_header(h) for h in raw_headers]
    header_map = dict(zip(raw_headers, normalized_headers, strict=False))

    rows: list[dict[str, str]] = []
    for row in reader:
        norm_row: dict[str, str] = {}
        for raw_h, value in row.items():
            norm_key = header_map.get(raw_h, "")
            if not norm_key:
                continue
            norm_row[norm_key] = str(value).strip() if value else ""
        rows.append(norm_row)
    return normalized_headers, rows


def _load_xlsx_rows(file: FileStorage) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "openpyxl is required for XLSX import. Install it and try again."
        ) from exc

    file.stream.seek(0)
    workbook = load_workbook(
        filename=io.BytesIO(file.stream.read()), read_only=True, data_only=True
    )
    sheet = workbook.active
    if sheet is None:
        raise ValueError("XLSX file has no active sheet.")

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        raw = next(rows_iter)
    except StopIteration:
        raise ValueError("XLSX file has no rows.") from None

    headers = [normalize_header(str(h)) for h in raw]

    rows: list[dict[str, str]] = []
    for row in rows_iter:
        norm_row: dict[str, str] = {}
        for idx, value in enumerate(row):
            h = headers[idx] if idx < len(headers) else ""
            if not h:
                continue
            norm_row[h] = str(value).strip() if value is not None else ""
        rows.append(norm_row)
    return headers, rows


def _load_xls_rows(file: FileStorage) -> tuple[list[str], list[dict[str, str]]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError(
            "xlrd is required for XLS import. Install it and try again."
        ) from exc

    file.stream.seek(0)
    workbook = xlrd.open_workbook(file_contents=file.stream.read())
    if workbook.nsheets == 0:
        raise ValueError("XLS file has no sheets.")
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("XLS file has no rows.")

    headers = [normalize_header(str(h)) for h in sheet.row_values(0)]

    rows: list[dict[str, str]] = []
    for row_idx in range(1, sheet.nrows):
        norm_row: dict[str, str] = {}
        for idx, value in enumerate(sheet.row_values(row_idx)):
            h = headers[idx] if idx < len(headers) else ""
            if not h:
                continue
            norm_row[h] = str(value).strip() if value else ""
        rows.append(norm_row)
    return headers, rows


def load_import_headers(file: FileStorage, fmt: str | None) -> list[tuple[str, str]]:
    """Return list of (normalized_key, display_label) tuples from file headers."""
    hint = fmt or (file.filename or "").rsplit(".", 1)[-1].lower()
    if hint == "csv":
        file.stream.seek(0)
        content = file.stream.read()
        file.stream.seek(0)
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        try:
            raw_headers = next(reader)
        except StopIteration:
            raise ValueError("CSV file has no rows.") from None
        normalized = [normalize_header(h) for h in raw_headers]
        return [
            (normalized[i], str(raw_headers[i]).strip())
            for i in range(len(raw_headers))
            if normalized[i]
        ]

    if hint == "xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("openpyxl is required for XLSX import.") from exc
        file.stream.seek(0)
        wb = load_workbook(
            filename=io.BytesIO(file.stream.read()), read_only=True, data_only=True
        )
        sheet = wb.active
        rows_iter = sheet.iter_rows(values_only=True)  # type: ignore[union-attr]
        try:
            raw = next(rows_iter)
        except StopIteration:
            raise ValueError("XLSX file has no rows.") from None
        raw_headers = [str(h).strip() if h is not None else "" for h in raw]
        normalized = [normalize_header(h) for h in raw_headers]
        return [
            (normalized[i], raw_headers[i])
            for i in range(len(raw_headers))
            if normalized[i]
        ]

    if hint == "xls":
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("xlrd is required for XLS import.") from exc
        file.stream.seek(0)
        wb = xlrd.open_workbook(file_contents=file.stream.read())
        if wb.nsheets == 0:
            raise ValueError("XLS file has no sheets.")
        sheet = wb.sheet_by_index(0)
        if sheet.nrows == 0:
            raise ValueError("XLS file has no rows.")
        raw_headers = [
            str(h).strip() if h is not None else "" for h in sheet.row_values(0)
        ]
        normalized = [normalize_header(h) for h in raw_headers]
        return [
            (normalized[i], raw_headers[i])
            for i in range(len(raw_headers))
            if normalized[i]
        ]

    raise ValueError(f"Unsupported format '{hint}'. Use csv, xlsx, or xls.")


def load_import_rows(
    file: FileStorage,
    fmt: str | None,
    required_headers: list[str],
    mapping: dict[str, str] | None = None,
) -> list[dict[str, str]]:
    """Load rows from a csv/xlsx/xls file, optionally applying a column mapping."""
    hint = fmt or (file.filename or "").rsplit(".", 1)[-1].lower()
    if hint == "csv":
        headers, rows = _load_csv_rows(file)
    elif hint == "xlsx":
        headers, rows = _load_xlsx_rows(file)
    elif hint == "xls":
        headers, rows = _load_xls_rows(file)
    else:
        raise ValueError(f"Unsupported format '{hint}'. Use csv, xlsx, or xls.")

    if mapping:
        missing_map = [h for h in required_headers if not mapping.get(h)]
        if missing_map:
            raise ValueError("Mapping missing for: " + ", ".join(missing_map))
        used = list(mapping.values())
        if len(set(used)) != len(used):
            raise ValueError("Mapping contains duplicate column selections.")
        unknown = [h for h in used if h not in headers]
        if unknown:
            raise ValueError(
                "Mapping references unknown headers: " + ", ".join(unknown)
            )
        return [
            {key: row.get(mapping[key], "").strip() for key in required_headers}
            for row in rows
        ]

    missing = [h for h in required_headers if h not in headers]
    if missing:
        raise ValueError("Missing required headers: " + ", ".join(missing))
    return rows


def save_import_file(file: FileStorage, prefix: str) -> tuple[str, Path, str]:
    """Save uploaded file to uploads/imports/<prefix>_<token>.<ext>. Returns (token, path, ext)."""
    filename = file.filename or "import"
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else "data"
    token = uuid4().hex
    import_dir = Path(current_app.root_path) / "uploads" / "imports"
    import_dir.mkdir(parents=True, exist_ok=True)
    file_path = import_dir / f"{prefix}_{token}.{extension}"
    file.save(file_path)
    return token, file_path, extension
