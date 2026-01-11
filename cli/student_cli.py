"""
Student management CLI tool.

This module provides command-line interface for managing student records,
including adding, updating, listing, and deleting students.
"""

import argparse
import csv
import logging
import sys
from pathlib import Path

from sqlalchemy.exc import SQLAlchemyError

from app import create_app
from app.models.student import Student, validate_email
from app.services.student_service import StudentService

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

REQUIRED_HEADERS = ["first_name", "last_name", "student_id", "email", "program"]


def _normalize_header(header: str | None) -> str:
    if not header:
        return ""
    return header.strip().lower()


def _load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as csv_file:
        reader = csv.DictReader(csv_file)
        raw_headers = reader.fieldnames or []
        normalized_headers = [_normalize_header(h) for h in raw_headers]
        header_map = dict(zip(raw_headers, normalized_headers, strict=False))
        missing = [h for h in REQUIRED_HEADERS if h not in normalized_headers]
        if missing:
            raise ValueError("Missing required headers: " + ", ".join(missing))

        rows: list[dict[str, str]] = []
        for row in reader:
            normalized_row: dict[str, str] = {}
            for raw_header, value in row.items():
                normalized_key = header_map.get(raw_header, "")
                if not normalized_key:
                    continue
                normalized_row[normalized_key] = str(value).strip() if value else ""
            rows.append(normalized_row)
        return rows


def _load_xlsx_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "openpyxl is required for XLSX import. Install it and try again."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise ValueError("XLSX file has no rows.") from None

    raw_headers = [_normalize_header(str(h)) for h in headers]
    missing = [h for h in REQUIRED_HEADERS if h not in raw_headers]
    if missing:
        raise ValueError("Missing required headers: " + ", ".join(missing))

    rows: list[dict[str, str]] = []
    for row in rows_iter:
        normalized_row: dict[str, str] = {}
        for idx, value in enumerate(row):
            header = raw_headers[idx] if idx < len(raw_headers) else ""
            if not header:
                continue
            normalized_row[header] = str(value).strip() if value is not None else ""
        rows.append(normalized_row)
    return rows


def _load_xls_rows(path: Path) -> list[dict[str, str]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError(
            "xlrd is required for XLS import. Install it and try again."
        ) from exc

    workbook = xlrd.open_workbook(path)
    if workbook.nsheets == 0:
        raise ValueError("XLS file has no sheets.")

    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("XLS file has no rows.")

    raw_headers = [_normalize_header(str(h)) for h in sheet.row_values(0)]
    missing = [h for h in REQUIRED_HEADERS if h not in raw_headers]
    if missing:
        raise ValueError("Missing required headers: " + ", ".join(missing))

    rows: list[dict[str, str]] = []
    for row_idx in range(1, sheet.nrows):
        row = sheet.row_values(row_idx)
        normalized_row: dict[str, str] = {}
        for idx, value in enumerate(row):
            header = raw_headers[idx] if idx < len(raw_headers) else ""
            if not header:
                continue
            normalized_row[header] = str(value).strip() if value else ""
        rows.append(normalized_row)
    return rows


def _load_rows(path: Path, file_format: str | None) -> list[dict[str, str]]:
    fmt = file_format or path.suffix.lstrip(".").lower()
    if fmt == "csv":
        return _load_csv_rows(path)
    if fmt == "xlsx":
        return _load_xlsx_rows(path)
    if fmt == "xls":
        return _load_xls_rows(path)
    raise ValueError(f"Unsupported format '{fmt}'. Use csv, xlsx, or xls.")


def _import_students(
    service: StudentService,
    rows: list[dict[str, str]],
    on_duplicate: str,
) -> int:
    created = updated = skipped = errors = 0
    seen_student_ids: set[str] = set()
    seen_emails: set[str] = set()

    for idx, row in enumerate(rows, start=2):
        missing_values = [header for header in REQUIRED_HEADERS if not row.get(header)]
        if missing_values:
            logger.error("Row %d missing values: %s", idx, ", ".join(missing_values))
            errors += 1
            continue

        student_id = row["student_id"].strip()
        email = row["email"].strip().lower()

        if not student_id:
            logger.error("Row %d missing student_id", idx)
            errors += 1
            continue
        if not validate_email(email):
            logger.error("Row %d invalid email: %s", idx, email)
            errors += 1
            continue

        if student_id in seen_student_ids or email in seen_emails:
            logger.error("Row %d duplicate within import file.", idx)
            errors += 1
            continue
        seen_student_ids.add(student_id)
        seen_emails.add(email)

        existing_by_id = service.query(Student).filter_by(student_id=student_id).first()
        existing_by_email = service.query(Student).filter_by(email=email).first()

        existing: Student | None = None
        if existing_by_id and existing_by_email:
            if existing_by_id.id != existing_by_email.id:
                logger.error(
                    "Row %d student_id/email belong to different records.", idx
                )
                errors += 1
                continue
            existing = existing_by_id
        else:
            existing = existing_by_id or existing_by_email

        if existing:
            if on_duplicate == "skip":
                skipped += 1
                continue
            if on_duplicate == "error":
                logger.error("Row %d duplicate found.", idx)
                errors += 1
                continue
            if on_duplicate == "update":
                updated_student = service.update_student(
                    existing.id,
                    row["first_name"],
                    row["last_name"],
                    student_id,
                    email,
                    row["program"],
                    validate_id=False,
                )
                if updated_student:
                    updated += 1
                else:
                    errors += 1
                continue

        try:
            service.add_student(
                row["first_name"],
                row["last_name"],
                student_id,
                email,
                row["program"],
                validate_id=False,
            )
            created += 1
        except ValueError as exc:
            logger.error("Row %d error: %s", idx, exc)
            errors += 1
        except SQLAlchemyError as exc:
            logger.error("Row %d database error: %s", idx, exc)
            errors += 1

    print(
        f"Import complete. Created: {created}, Updated: {updated}, "
        f"Skipped: {skipped}, Errors: {errors}"
    )
    return 0 if errors == 0 else 1


def main() -> int:
    """
    Main CLI entry point.

    Returns:
        Exit code (0 for success, 1 for error)
    """
    parser = argparse.ArgumentParser(
        description="Student management CLI tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    subparsers = parser.add_subparsers(dest="command", help="Available commands")

    # Add command
    add_parser = subparsers.add_parser("add", help="Add a new student")
    add_parser.add_argument("--first-name", required=True, help="First name")
    add_parser.add_argument("--last-name", required=True, help="Last name")
    add_parser.add_argument(
        "--student-id",
        required=True,
        help="Matrikelnummer (8 digits)",
    )
    add_parser.add_argument("--email", required=True, help="Email address")
    add_parser.add_argument("--program", required=True, help="Study program/major")

    # List command
    list_parser = subparsers.add_parser("list", help="List all students")
    list_parser.add_argument("--search", help="Search by name, student ID, or email")
    list_parser.add_argument("--program", help="Filter by program")

    # Show command
    show_parser = subparsers.add_parser("show", help="Show student details")
    show_parser.add_argument("id", type=int, help="Database ID")

    # Update command
    update_parser = subparsers.add_parser("update", help="Update a student")
    update_parser.add_argument("id", type=int, help="Database ID")
    update_parser.add_argument("--first-name", help="New first name")
    update_parser.add_argument("--last-name", help="New last name")
    update_parser.add_argument("--student-id", help="New student ID")
    update_parser.add_argument("--email", help="New email")
    update_parser.add_argument("--program", help="New program")

    # Delete command
    delete_parser = subparsers.add_parser("delete", help="Delete a student")
    delete_parser.add_argument("id", type=int, help="Database ID")

    # Import command
    import_parser = subparsers.add_parser(
        "import", help="Import students from CSV/XLSX"
    )
    import_parser.add_argument(
        "--file",
        "-f",
        required=True,
        help="Path to CSV or XLSX file",
    )
    import_parser.add_argument(
        "--format",
        choices=["csv", "xlsx", "xls"],
        help="Optional file format override",
    )
    import_parser.add_argument(
        "--on-duplicate",
        choices=["skip", "update", "error"],
        default="skip",
        help="How to handle duplicates (default: skip)",
    )

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return 1

    # Create Flask app context for database access
    app = create_app()

    with app.app_context():
        # Initialize service
        service = StudentService()

        try:
            if args.command == "add":
                student = service.add_student(
                    args.first_name,
                    args.last_name,
                    args.student_id,
                    args.email,
                    args.program,
                )
                print(
                    f"Created student: DB-ID={student.id}, "
                    f"Name={student.first_name} {student.last_name}, "
                    f"Matrikelnummer={student.student_id}, "
                    f"Email={student.email}"
                )
                return 0

            if args.command == "list":
                students = service.list_students(args.search, args.program)
                if students:
                    print(f"\nFound {len(students)} students:\n")
                    print(
                        f"{'DB ID':<7} {'Matrikelnr.':<12} {'Name':<35} {'Email':<35} {'Program':<30}"
                    )
                    print("-" * 125)
                    for student in students:
                        name = f"{student.first_name} {student.last_name}"
                        print(
                            f"{student.id:<7} {student.student_id:<12} {name:<35} {student.email:<35} {student.program:<30}"
                        )
                else:
                    print("No students found")
                return 0

            if args.command == "show":
                student = service.get_student(args.id)
                if student:
                    print("\nStudent Details:")
                    print(f"  Datenbank-ID: {student.id}")
                    print(f"  Matrikelnummer: {student.student_id}")
                    print(f"  Name: {student.first_name} {student.last_name}")
                    print(f"  Email: {student.email}")
                    print(f"  Program: {student.program}")
                    print(f"  Created: {student.created_at}")
                    print(f"  Updated: {student.updated_at}")
                else:
                    print(f"Student mit DB-ID {args.id} nicht gefunden")
                    return 1
                return 0

            if args.command == "update":
                student = service.update_student(
                    args.id,
                    args.first_name,
                    args.last_name,
                    args.student_id,
                    args.email,
                    args.program,
                )
                if student:
                    print(
                        f"Updated student: DB-ID={student.id}, "
                        f"Name={student.first_name} {student.last_name}, "
                        f"Matrikelnummer={student.student_id}, "
                        f"Email={student.email}"
                    )
                    return 0
                print(f"Student mit DB-ID {args.id} nicht gefunden")
                return 1

            if args.command == "delete":
                if service.delete_student(args.id):
                    print(f"Student mit DB-ID {args.id} erfolgreich gelöscht")
                    return 0
                print(f"Student mit DB-ID {args.id} nicht gefunden")
                return 1

            if args.command == "import":
                path = Path(args.file)
                if not path.exists():
                    print(f"File not found: {path}", file=sys.stderr)
                    return 1
                rows = _load_rows(path, args.format)
                return _import_students(service, rows, args.on_duplicate)

        except ValueError as e:
            logger.error(f"Validation error: {e}")
            print(f"Error: {e}", file=sys.stderr)
            return 1

        except SQLAlchemyError as e:
            logger.error(f"Database error: {e}", exc_info=True)
            print("Database error. Please try again.", file=sys.stderr)
            return 1

        except KeyboardInterrupt:
            logger.info("Operation cancelled by user")
            print("\nOperation cancelled.", file=sys.stderr)
            return 130

        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            print(f"Unexpected error: {e}", file=sys.stderr)
            return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
